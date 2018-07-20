import xlsxwriter
import os, fnmatch
import xlrd
from openpyxl.workbook import Workbook


def xlsxwrite(header, data, name):
    header = [header]
    new_data = data

    #wb = Workbook(write_only=True)
    #dest_filename = name + '.xlsx'
    #ws1 = wb.create_sheet()
    #ws1.title = name
    #ws1.append(header)

    #for row in new_data:
    #    ws1.append(row)

    #wb.save(filename = dest_filename)

    workbook = xlsxwriter.Workbook(name + '.xlsx')
    worksheet = workbook.add_worksheet(name)
    table = header
    table += new_data
    for row_idx, row in enumerate(table):
        for col_idx, item in enumerate(row):
            worksheet.write(row_idx, col_idx, item)
    workbook.close()

def readxlsx(pathname):

    path = pathname

    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_by_index(0)
    # Change this depending on how many header rows are present
    # Set to 0 if you want to include the header data.
    offset = 1

    rows = []
    for i, row in enumerate(range(worksheet.nrows)):
        if i <= offset-1:  # (Optionally) skip headers
            continue
        r = []
        for j, col in enumerate(range(worksheet.ncols)):
            r.append(worksheet.cell_value(i, j))
        rows.append(r)

    print('Got ',len(rows) - offset,' rows')
    #print(rows[0])  # Print first data row sample
    #print(rows[offset])
    return rows


def find_files(directory, pattern):
    for root, dirs, files in os.walk(directory):
        for basename in files:
            if fnmatch.fnmatch(basename, pattern):
                filename = os.path.join(root, basename)
                yield filename


def makenewpersons(rows, oldorders, oldpersons, plist, er):
    newpersons = []
    er = [['Группа'], ['Строка с содержимым'], ['Ошибка']]
    for row in rows:
        name = (' '.join(row[1].split())).split(' ', 2)
        if (row[0] if isinstance(row[0], str) else str(int(float(row[0])))) in er[1]:
            continue
        if name[0] == '':
            print(row)
            if len(row)>12 and row[12] != '':
                er[0] += [row[12]]
                er[1] += [(row[0] if isinstance(row[0], str) else str(int(float(row[0]))))]
                er[2] += ['Нет имени студента']
            continue
        if row[1] in er[1]:
            continue
        if '(' in row[1]:
            er[0] += [row[12]]
            er[1] += [row[1]]
            er[2] += ['Скобки в имени студента?']
            continue
        if len(name)>2 and name[1] == '':
            name = [name[0], name[2].split(' ', 1)[0], name[2].split(' ', 1)[1] if len(name[2].split(' ', 1))>1 else '']
        if len(name) == 2:
            er[0] += [row[12]]
            er[1] += [row[1]]
            er[2] += ['Нет отчества? Пробелы?']
            continue
        if row[1] not in plist[1]:
            plist[0] += [(row[0] if isinstance(row[0], str) else str(int(float(row[0]))))]
            plist[1] += [row[1]]
            plist[2] += [name]
    print(plist[0])
    print(plist[1])
    print(plist[2])
    print(rows[0])
    print(len(plist[0]), len(plist[1]), len(plist[2]))

    pn = 0
    nonen = 0
    check = False
    for p in plist[0]:
        if p != '':
            for row in oldorders:
                if (p if isinstance(p, str) else str(int(float(p)))) == str(row[1]):
                    #print(p, ' ', str(row[1]))
                    plist[3] += [row[0]]
                    #print(plist[3][pn], ' ', [row[0]])
                    check = True
                    break
        else:
            nonen += 1
            plist[3] += ['none']
            check = True
        if not check:
            nonen += 1
            plist[3] += ['none']
        check = False
        pn += 1
    print(plist[0])
    print(plist[3])
    print(len(plist[0]), len(plist[1]), len(plist[2]), len(plist[3]), nonen)

    pn = 0
    nonen = 0
    check = False
    for p in plist[2]:
        if plist[3][pn] == 'none':
            for row in oldpersons:
                #print('----------------')
                #print(p)
                #print(p[0].strip())
                #print(row[1].strip())
                #print(p[1].strip())
                #print(row[2].strip())
                #print(p[2].strip())
                #print(row[3].strip())
                if p[0].strip() == row[1].strip() and p[1].strip() == row[2].strip() and p[2].strip() == row[3].strip():
                    plist[3][pn] = row[0]
                    check = True
            if not check:
                nonen += 1
            check = False
        pn += 1
    print(plist[2])
    print(plist[3])
    print(len(plist[0]), len(plist[1]), len(plist[2]), len(plist[3]), nonen)

    pn = 0
    nonen = 0
    check = False
    for p in plist[2]:
        if plist[3][pn] == 'none':
            for row in oldpersons:
                #print('----------------')
                #print(p)
                #print(p[0].strip())
                #print(row[1].strip())
                #print(p[1].strip())
                #print(row[2].strip())
                #print(p[2].strip())
                #print(row[3].strip())
                if p[0].strip() == row[1].strip() and p[1].strip()[0] == row[2].strip()[0] and p[2].strip()[0] == row[3].strip()[0]:
                    plist[3][pn] = row[0]
                    check = True
            if not check:
                nonen += 1
            check = False
        pn += 1
    print(plist[2])
    print(plist[3])
    print(len(plist[0]), len(plist[1]), len(plist[2]), len(plist[3]), nonen)

    pn = 0
    nonen = 0
    check = False
    print(plist[2])
    print(plist[3])
    newp = []
    for p in plist[2]:
        p += [(plist[0][pn] if isinstance(plist[0][pn], str) else str(int(float(plist[0][pn]))))]
        p += ['РФ']
        p += ['33301']
        p += [plist[1][pn]]
        if plist[3][pn] != 'none':
            p += [str(int(float(plist[3][pn])))]
        else:
            p += [str(pn + 1000)]
            newp += [p]
            if plist[3][pn] == 'none':
                #print(plist[0][pn])
                print(plist[2][pn])
                #print(plist[3][pn])
        pn += 1
    print(plist[2])
    print(plist[3])
    ferrors = []
    n = 0
    for r in er[0]:
        ferrors += [[er[0][n], er[1][n], er[2][n]]]
        n += 1
    print(ferrors)
    return newp, plist[2], ferrors


def makeneworders(allpersons, rows, er):
    # 0-Фамилия, 1-Имя, 2-Отчество, 3-Зачетка, 4-Гражданство, 5-ДатаРождения, 6-ФИО, 7-ID
    # 0-'№ зачетки', 1-'Студент (ФИО)', 2-'Тип ведомости', 3-'Вид ведомости', 4-'№ ведомости', 5-'Учебный год', 6-'Институт', 7-'Форма обучения', 8-'Уровень подготовки', 9-'Шифр направления подготовки', 10-'Профиль\Специализация', 11-'Курс', 12-'Группа', 13-'Дисциплина', 14-'Период контроля', 15-'Вид занятия', 16-'Система оценивания', 17-'Отметка', 18-'Дата занятия'
    ptuple = []
    orders = []
    pn = 0
    for p in allpersons:
        for row in rows:
            if p[6] == row[1]:
                if (not isinstance(row[11], float) and not isinstance(row[11], int)) and row[11] in er[-1][1]:
                    continue
                if not (isinstance(row[11], float) or isinstance(row[11], int)):
                    er += [[row[11], row[11], 'В курсе указано не целое число? Подозрение на неправильный порядок колонок!']]
                    continue
                if [row[1], int(float(row[11])), str(row[12]).strip(' ')] not in ptuple: # + 9, 10, 7, 5
                    ptuple += [[row[1], int(float(row[11])), str(row[12]).strip(' ')]]
                    if isinstance(row[12], float):
                        continue
                    orders += [[row[1], p[7], p[3], int(float(row[11])), str(row[12]).strip(' '), row[10], str(row[9]).strip(' '), row[7].strip(' '), '201'+str(int(row[12].split('-', 1)[1][0])+int(float(row[11]))-1)+'-201'+str(int(row[12].split('-', 1)[1][0])+int(float(row[11])))]]
                    pn += 1
    neworders = []
    print(len(ptuple), len(orders))
    for tuple in ptuple:
        print(tuple)
    pn = 0
    for order in orders:
        print(order)
        pn += 1
        if order[4] in ['2-52-44.06.01', '2-52-44.06.02', '2-52-44.06.03', '2-52-44.06.04', '2-52-44.06.05', '2-52-44.06.06', '2-52-44.06.07', '2-52-44.06.08', '2-52-44.06.09', '2-52-44.06.10', '2-52-44.06.11', '2-52-44.06.12', '2-52-44.06.13', '2-52-44.06.14', '2-52-44.06.15']:
            order[4] = '2-52-44.06.01'
        if order[4] in ['201-51', '301-51', '401-51', '601-51', '701-51', '801-51', '901-51', '1001-51', '1101-51', '1201-51', '1301-51', '1401-51']:
            order[4] = '101-51'
        if order[4] in ['201-71', '301-71', '401-71', '601-71', '701-71']:
            order[4] = '101-71'
        if order[4] == '40351':
            order[4] = '403-51'
        if order[4] in ['4-73-38.06.02', '4-73-38.06.03', '4-73-38.06.04', '4-73-38.06.05', '4-73-38.06.06', '4-73-38.06.07', '4-73-38.06.08', '4-73-38.06.09', '4-73-38.06.10', '4-73-38.06.11', '4-73-38.06.12', '4-73-38.06.13', '4-73-38.06.14']:
            order[4] = '4-73-38.06.01'

        neworders += [[('о' if order[6].split('.', 2)[1]=='08' else ('а' if order[6].split('.', 2)[1]=='06' else '')) + order[1],
                       order[2],
                       order[4]+'-'+str(order[3])+'-c',
                       '01.09.'+order[8].split('-', 1)[0],
                       ('Зачисление в вуз' if order[3] == 1 else 'Перевод на следующий курс') if (order[6].split('.', 2)[1]=='03' or order[6].split('.', 2)[1]=='04') else ('Зачисление в аспирантуру' if order[3] == 1 else 'Перевод на следующий курс аспирантуры'),
                        '',
                        '01.09.'+order[8].split('-', 1)[0],
                       '30.06.'+order[8].split('-', 1)[1],
                        order[6],
                        order[7],
                        order[8],
                        order[3],
                        'Бюджет',
                        ''.join(order[4].split()),
                        '',
                        '',
                        order[5]]]
    print(er)
    for neworder in neworders:
        print(neworder)
    print(len(neworders))
    # 0-'IDФизическогоЛица', 1-'№ЗачетнойКнижки', 2-'КанцНомерПриказа', 3-'КанцДатаПриказа', 4-'ВидПриказа', 5-'ЗаголовокПриказа', 6-'ДатаНачала', 7-'ДатаОкончания', 8-'КодНаправленияПодготовки', 9-'ФормаОбучения', 10-'УчебныйГод', 11-'Курс', 12-'ОснованиеОбучения', 13-'УчебнаяГруппа', 14-'УчебнаяПодгруппа', 15-'Аналитика', 16-'Профиль\Специализация'
    return neworders, er


rows = []

for filename in find_files('C:\\Users\\HYPER\\PycharmProjects\\autogenerateddata', '*.xlsx'):
    print('Found xlsx source:', filename)
    rows += readxlsx(filename)

for filename in find_files('C:\\Users\\HYPER\\PycharmProjects\\autogenerateddata', '*.xls'):
    print('Found xlsx source:', filename)
    rows += readxlsx(filename)

print('stolen ',len(rows), 'rows')
#print(rows[0])

register = []
name = []
tuple = []
id = []
plist = [register, name, tuple, id]
oldpersons = []
oldorders = []
er = []
newpersons = []
allpersons = []

for filename in find_files('C:\\Users\\HYPER\\Desktop\\Работа', 'Физические лица.xls'):
    print('Found xlsx source:', filename)
    oldpersons += readxlsx(filename)
for filename in find_files('C:\\Users\\HYPER\\Desktop\\Работа', 'Физические лица - аспиранты.xls'):
    print('Found xlsx source:', filename)
    oldpersons += readxlsx(filename)
for filename in find_files('C:\\Users\\HYPER\\Desktop\\Работа', 'Физические лица  Ординаторы 1 год  (2017-2019).xls'):
    print('Found xlsx source:', filename)
    oldpersons += readxlsx(filename)
for filename in find_files('C:\\Users\\HYPER\\Desktop\\Работа', 'Физические лица  Ординаторы 2 год  (2016-2018) выпуск.xls'):
    print('Found xlsx source:', filename)
    oldpersons += readxlsx(filename)
for filename in find_files('C:\\Users\\HYPER\\Desktop\\Работа', 'Физические лица доп1.xlsx'):
    print('Found xlsx source:', filename)
    oldpersons += readxlsx(filename)

for filename in find_files('C:\\Users\\HYPER\\Desktop\\Работа', 'Приказы.xls'):
    print('Found xlsx source:', filename)
    oldorders += readxlsx(filename)

print(oldpersons[0])
print(oldorders[0])

name = 'Физические лица'
# 0-Фамилия, 1-Имя, 2-Отчество, 3-Зачетка, 4-Гражданство, 5-ДатаРождения, 6-ФИО, 7-ID
pers, allpersons, er = makenewpersons(rows, oldorders, oldpersons, plist, er)
for row in pers:
    newpersons += [[row[7], row[0], row[1], row[2], '', '', float(row[5]), '', row[4]]]
header = ['']
#xlsxwrite(header, newpersons, name)

name = 'ГиперВедомость'
header = ['№ зачетки', 'Студент (ФИО)', 'Тип ведомости', 'Вид ведомости', '№ ведомости', 'Учебный год', 'Институт', 'Форма обучения', 'Уровень подготовки', 'Шифр направления подготовки', 'Профиль\Специализация', 'Курс', 'Группа', 'Дисциплина', 'Период контроля', 'Вид занятия', 'Система оценивания', 'Отметка', 'Дата занятия']
#xlsxwrite(header, rows, name)

name = 'Приказы'
neworders, er = makeneworders(allpersons, rows, er)
header = ['IDФизическогоЛица', '№ЗачетнойКнижки', 'КанцНомерПриказа', 'КанцДатаПриказа', 'ВидПриказа', 'ЗаголовокПриказа', 'ДатаНачала', 'ДатаОкончания', 'КодНаправленияПодготовки', 'ФормаОбучения', 'УчебныйГод', 'Курс', 'ОснованиеОбучения', 'УчебнаяГруппа', 'УчебнаяПодгруппа', 'Аналитика', 'Профиль\Специализация']
xlsxwrite(header, neworders, name)

